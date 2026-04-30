"""매핑 안 된 후기 태그를 엑셀로 export"""
import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout.reconfigure(encoding='utf-8')

with open('reviews_data.json', encoding='utf-8') as f:
    d = json.load(f)

# 이미 Q-옵션 ID 에 매핑된 태그
MAPPED = {
    '전업수험생','직장병행','재학수험생','육아병행','퇴직수험생','아르바이트병행',
    '국가직9급','지방직9급','서울시9급',
    '행정직','세무직','교육직','기술직','간호보건직',
    '교정직','검찰직','보호직','철도경찰직','출입국관리직',
    '노베이스','수능베이스','전공자','토익베이스',
    '6개월미만','6개월~1년','1년~1년6개월','1년6개월~2년','2년~3년',
    '하루10시간미만','하루10시간이상',
    '인강병행','기출반복','회독반복','독학위주','노량진현강','스터디병행',
    '멘탈관리','슬럼프극복','영어약자','효율중심','이해중심학습','커리큘럼중심',
}

# 그룹 분류
GROUPS = {
    'A. 학습법 디테일 (Q7 보완)': [
        '하프모의고사','동형모의고사','단권화','오답노트','빈출위주','매일전과목',
        '기본서회독','발췌독활용','압축이론','두문자어암기','영어단어매일암기',
        '스토리텔링암기','형광펜활용','무한반복','5급기출활용','암기중심',
        '복습중심','문제풀이중심','판례중심','선지분석','빈출위주','전략적시험운영',
        '노트정리','이론중심','이미지암기','키워드암기','OX중심','맵핑학습',
        '기출위주','기출반복','회독반복','독학위주','요약본제작','말노트활용',
        '필기노트','손필기노트','마인드맵활용','서브노트제작','단어장회독','단어장암기',
        '백지공부법','WOX회독법','과목깨기','5분발표','마킹연습','선지분석',
        '소거법활용','문법위주','독해위주','이해중심학습','암기중심학습','개념중심',
        '아웃풋중심','질적학습','공식암기','오답풀이','막판스퍼트','벼락치기',
        '영단어매일암기','취침전단어암기','오전암기','약점공략','약점보강','오답분석',
        '오답복기','약점공략','강의노트활용','법령집활용','대학교재활용','모의문제제작',
        '모의고사활용','블로그정리','타이핑단권화','기화펜활용','한자자격증','조각공부',
    ],
    'B. 시간대·루틴 (Q6 보완)': [
        '규칙적루틴','아침형','주1일휴식','무휴일공부','저녁형','새벽공부',
        '자투리시간활용','이동시간암기','식사중학습','취침전단어암기','비문학아침풀기',
        '주2일휴식','주5일공부','주6일공부','하루14시간이상','하루10시간',
        '불규칙루틴','하루단위목표','꾸준함','낮잠루틴','주말집중',
        '시간관리','수면관리','컨디션관리','체력관리','스트레스관리','식단관리',
        '서서공부','휴대폰차단','커뮤니티차단','오픈채팅활용','40대','30대초반',
        '30대중반','30대후반','20대초반','시간단위목표','막판스퍼트',
    ],
    'C. 학습 환경': [
        '스터디카페','독서실공부','도서관공부','관리형독서실','집공부','카페공부','캠스터디',
    ],
    'D. 학습 도구·콘텐츠': [
        '앱활용','유튜브활용','무료강의활용','AI활용','마인드맵활용','아이패드활용',
        '손필기노트','타이머활용','포스트잇활용','법령집활용','단어장회독','노트정리',
        '서브노트제작','백지공부법','WOX회복법','말노트활용','대학교재활용','5급기출활용','휴대폰차단',
        '커뮤니티차단','오픈채팅활용','단어장암기','단원별모의고사','블로그정리','요약본제작',
    ],
    'E. 응시 횟수·이력 (Q1 보완)': [
        '초시생','재시생','재직수험생','휴학수험생','복학수험생','전역수험생',
        '군복무수험생','통학수험생','해외거주수험생','워킹맘','맘시생','경력단절',
        '사업경험','전직렬전환','학업병행','3년이상','장기수험생','경력직','2년제',
        '기혼','지방직병행','7급병행','국가직병행','군무원병행','면접병행','복학수험생',
        '면접독학','PSAT병행','1차2차병행','공기업병행','현강병행','집공부',
    ],
    'F. 베이스·전공 (Q4 보완)': [
        '비전공자','이과출신','이공계출신','회계세무전공','한능검베이스','영어베이스',
        '법학전공','경영학전공','형사소송법베이스','한국사베이스','임용베이스',
        '관세사베이스','영어강사경력','사범대','기사자격증','한자자격증','문과출신',
        '국시베이스','세무사베이스','재경관리사베이스','행정법베이스','행정학전공',
        '역사학전공','사범대','국어전공','영어전공','국어베이스','경찰직베이스',
        '제2외국어','전기기사','지텔프','한능검','전공위주','검찰직베이스','시간관리',
    ],
    'G. 시험·면접 (Q2 보완)': [
        '국가직7급','지방직7급','서울시7급','지방직8급','군무원9급','군무원7급',
        '면접병행','국가직병행','지방직병행','PSAT병행','1차2차병행','면접스터디',
        '면접독학','면접우수','모의면접','재면접합격','필합면탈경험','가산점합격',
        '재학수험생','7급병행','공기업병행','2관왕','국가직7급','학업병행',
    ],
    'H. 직렬 보완 (Q3 보완)': [
        '사회복지직','공업직','전산직','건축직','관세직','전기직','기계직',
        '농업직','임업직','토목직','시설직','외무영사직','우정행정직','계리직',
        '방송통신직','환경직','군수직','검찰직','경찰직','보호직','철도경찰직','지방직7급','기계일반','출입국관리직','계리직',
    ],
    'I. 기간 보완 (Q5 보완)': [
        '단기합격','단기완성','3년이상','장기수험생','막판스퍼트',
    ],
    'J. 걱정·관리 (Q8 보완)': [
        '한국사약자','국어약자','컨디션관리','수면관리','체력관리','스트레스관리',
        '시간관리','건강이슈','번아웃경험','과락경험','모의고사슬럼프','강사변경',
        '휴대폰차단','커뮤니티차단','자기암시','식단관리','강사변경','마음잡기',
    ],
    'K. 합격 결과·강조': [
        '고득점합격','단기합격','단기완성','가산점합격','2관왕','면접우수',
    ],
    'L. 인적 특성': [
        '40대','30대초반','30대중반','30대후반','20대초반','고졸수험생',
        '4년제','지방4년제','서울4년제','2년제','지거국','지역인재전형',
        '장애인전형','기혼','경력직','임용수험생',
    ],
    'M. 강사·강의 특수': [
        '2배속수강','1.5배속수강','1.2배속수강','3배속수강','완강중심',
        '강사변경','현강병행','영어강사경력',
    ],
}

# 모든 매핑 안 된 태그 (빈도 정렬)
all_tags = d['tagCounts']
unmapped = sorted(
    [(t, c) for t, c in all_tags.items() if t not in MAPPED],
    key=lambda x: -x[1]
)

# 태그 → 그룹 역매핑 (앞순서 그룹 우선)
tag_to_group = {}
for group_name, tags in GROUPS.items():
    for t in tags:
        if t not in tag_to_group and t not in MAPPED:
            tag_to_group[t] = group_name

# 엑셀 생성
wb = Workbook()

# 스타일
header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill('solid', fgColor='1A1A2E')
group_fill = PatternFill('solid', fgColor='F0EDE6')
border = Border(
    left=Side(style='thin', color='D0CEC7'),
    right=Side(style='thin', color='D0CEC7'),
    top=Side(style='thin', color='D0CEC7'),
    bottom=Side(style='thin', color='D0CEC7'),
)
center = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

# ──────────── 시트 1: 전체 (빈도순) ────────────
ws1 = wb.active
ws1.title = '전체 빈도순'

headers = ['순위', '태그명', '빈도', '제안 그룹', '비고']
for col, h in enumerate(headers, 1):
    c = ws1.cell(1, col, h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

for idx, (tag, cnt) in enumerate(unmapped, 1):
    ws1.cell(idx + 1, 1, idx).alignment = center
    ws1.cell(idx + 1, 2, tag).alignment = left_align
    ws1.cell(idx + 1, 3, cnt).alignment = center
    ws1.cell(idx + 1, 4, tag_to_group.get(tag, '미분류')).alignment = left_align
    ws1.cell(idx + 1, 5, '').alignment = left_align
    for col in range(1, 6):
        ws1.cell(idx + 1, col).border = border

ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 8
ws1.column_dimensions['D'].width = 30
ws1.column_dimensions['E'].width = 30
ws1.freeze_panes = 'A2'

# ──────────── 시트 2: 그룹별 ────────────
ws2 = wb.create_sheet('그룹별')

ws2.cell(1, 1, '그룹').font = header_font
ws2.cell(1, 2, '태그명').font = header_font
ws2.cell(1, 3, '빈도').font = header_font
ws2.cell(1, 4, '비고').font = header_font
for col in range(1, 5):
    ws2.cell(1, col).fill = header_fill
    ws2.cell(1, col).alignment = center
    ws2.cell(1, col).border = border

row = 2
unmapped_dict = dict(unmapped)
for group_name, tags in GROUPS.items():
    group_tags = []
    for t in tags:
        if t in unmapped_dict and t not in [x[0] for x in group_tags]:
            group_tags.append((t, unmapped_dict[t]))
    if not group_tags:
        continue
    group_tags.sort(key=lambda x: -x[1])

    # 그룹 헤더
    ws2.cell(row, 1, group_name).font = Font(bold=True, size=12)
    ws2.cell(row, 1).fill = group_fill
    for col in range(1, 5):
        ws2.cell(row, col).fill = group_fill
        ws2.cell(row, col).border = border
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    for tag, cnt in group_tags:
        ws2.cell(row, 1, '').alignment = center
        ws2.cell(row, 2, tag).alignment = left_align
        ws2.cell(row, 3, cnt).alignment = center
        ws2.cell(row, 4, '').alignment = left_align
        for col in range(1, 5):
            ws2.cell(row, col).border = border
        row += 1
    row += 1  # 그룹 사이 빈 행

ws2.column_dimensions['A'].width = 32
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 8
ws2.column_dimensions['D'].width = 30
ws2.freeze_panes = 'A2'

# ──────────── 시트 3: 미분류 only ────────────
unclassified = [(t, c) for t, c in unmapped if t not in tag_to_group]
ws3 = wb.create_sheet('미분류')

for col, h in enumerate(['순위', '태그명', '빈도', '비고'], 1):
    c = ws3.cell(1, col, h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border

for idx, (tag, cnt) in enumerate(unclassified, 1):
    ws3.cell(idx + 1, 1, idx).alignment = center
    ws3.cell(idx + 1, 2, tag).alignment = left_align
    ws3.cell(idx + 1, 3, cnt).alignment = center
    ws3.cell(idx + 1, 4, '').alignment = left_align
    for col in range(1, 5):
        ws3.cell(idx + 1, col).border = border

ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 22
ws3.column_dimensions['C'].width = 8
ws3.column_dimensions['D'].width = 30
ws3.freeze_panes = 'A2'

# 저장
out_path = '답변정리/매핑안된_후기태그.xlsx'
wb.save(out_path)
print(f'✓ 엑셀 저장됨: {out_path}')
print(f'  - 시트 1 (전체 빈도순): {len(unmapped)} 행')
print(f'  - 시트 2 (그룹별): {len(GROUPS)} 그룹')
print(f'  - 시트 3 (미분류): {len(unclassified)} 행')
